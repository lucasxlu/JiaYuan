import requests
from openpyxl import load_workbook

import os
# Python Image Library haven't been supported on Python3.X yet
# from PIL import Image
from io import BytesIO

from people.candidate import Candidate


def download_avatar(candidate, download_dir):
    image_dir = download_dir + '/' + str(candidate.gender) + '/' + str(candidate.province)
    if not os.path.exists(image_dir):
        os.makedirs(image_dir)

    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2882.4 "
                      "Safari/537.36",
        "Accept - Encoding": "gzip, deflate, sdch"
    }

    try:
        response = requests.get(candidate.image, headers=headers, timeout=30)
        if response.status_code == 200:
            with open(image_dir + '/' + str(candidate.uid) + '.' + candidate.image.split(".")[
                        len(candidate.image.split(".")) - 1], mode='wb') as f:
                f.write(response.content)
                f.flush()
                f.close()
        elif response.status_code == 403:
            print('Access Denied!')
        else:
            print(str(response.status_code) + ' ;URL is ' + candicate.image)
    except:
        print('Oops...There is sth wrong...')
        pass


def read_excel(excel_path):
    candidate_list = list()
    wb = load_workbook(excel_path)
    ws = wb.active
    for i in range(2, ws.max_row - 1):  # -1 means that the last row is null
        candidate = Candidate(uid=ws.cell(row=i, column=1).value, nickname=ws.cell(row=i, column=2).value,
                              age=ws.cell(row=i, column=3).value, height=ws.cell(row=i, column=4).value,
                              image=ws.cell(row=i, column=5).value, marriage=ws.cell(row=i, column=6).value,
                              education=ws.cell(row=i, column=7).value, work_location=ws.cell(row=i, column=8).value,
                              work_sublocation=ws.cell(row=i, column=9).value,
                              shortnote=ws.cell(row=i, column=10).value,
                              matchCondition=ws.cell(row=i, column=11).value,
                              randListTag=ws.cell(row=i, column=12).value,
                              province=ws.cell(row=i, column=13).value, gender=ws.cell(row=i, column=14).value)
        candidate_list.append(candidate)
        # print(candidate)

    return candidate_list


if __name__ == '__main__':
    candidate_list = read_excel("D:/JiaYuan_Excel/女性用户-重庆.xlsx")
    for candicate in candidate_list:
        download_avatar(candicate, 'd:/JiaYuan_Avatar')
