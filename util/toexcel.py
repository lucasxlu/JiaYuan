import json
import os
import logging
from bs4 import BeautifulSoup

from openpyxl import Workbook


def json_to_list(lover_root_dir):
    lover_lists_by_province = []  # 每一个特定省份的list
    for each_json_file in os.listdir(lover_root_dir):
        with open(lover_root_dir + '/' + each_json_file, 'r', encoding='utf-8') as f:
            logging.debug('正在处理 ' + each_json_file + '.....')
            for each_line in f.readlines():
                repaired_each_line = each_line.replace('\'', '\"').replace('None', 'null').replace('\\u', '').replace(
                    '\\xa0', '')
                json_content = '{"userInfo" :' + repaired_each_line + '}'
                print(json_content)
                try:
                    json_obj = json.loads(
                        json_content, encoding='utf-8')
                    print(json_obj)
                    lover_lists_by_province.append(json_obj)
                except:
                    pass

    return lover_lists_by_province


def write_excel(lists, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "用户信息"
    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=2).value = '姓名'
    ws.cell(row=1, column=3).value = '年龄'
    ws.cell(row=1, column=4).value = '身高'
    ws.cell(row=1, column=5).value = '头像'
    ws.cell(row=1, column=6).value = '婚姻状态'
    ws.cell(row=1, column=7).value = '教育程度'
    ws.cell(row=1, column=8).value = '工作地点1'
    ws.cell(row=1, column=9).value = '工作地点2'
    ws.cell(row=1, column=10).value = '简要介绍'
    ws.cell(row=1, column=11).value = '择偶标准'
    ws.cell(row=1, column=12).value = '用户标签'

    rownum = 2

    for each_item in lists:
        info_list = each_item.get('userInfo')
        for each_job_info_obj in info_list:
            # try:
            ws.cell(row=rownum, column=1).value = each_job_info_obj['uid']
            ws.cell(row=rownum, column=2).value = each_job_info_obj['nickname']
            ws.cell(row=rownum, column=3).value = each_job_info_obj['age']
            ws.cell(row=rownum, column=4).value = each_job_info_obj['height']
            ws.cell(row=rownum, column=5).value = each_job_info_obj['image']
            ws.cell(row=rownum, column=6).value = each_job_info_obj['marriage']
            ws.cell(row=rownum, column=7).value = each_job_info_obj['education']
            ws.cell(row=rownum, column=8).value = each_job_info_obj['work_location']
            ws.cell(row=rownum, column=9).value = each_job_info_obj['work_sublocation']
            ws.cell(row=rownum, column=10).value = each_job_info_obj['shortnote']
            ws.cell(row=rownum, column=11).value = each_job_info_obj['matchCondition']
            ws.cell(row=rownum, column=12).value = BeautifulSoup(each_job_info_obj['randListTag'],
                                                                 'html.parser').find('span').get_text()
            # except:
            #     pass
            rownum += 1
    wb.save('d:/' + filename + '.xlsx')
    logging.info('Excel生成成功!')


def process(jiayuan_data_base_root):
    for gender_dir in os.listdir(jiayuan_data_base_root):
        for province_dir in os.listdir(jiayuan_data_base_root + os.path.sep + gender_dir):
            user_lists = json_to_list(jiayuan_data_base_root + os.path.sep + gender_dir + os.path.sep + province_dir)
            excel_name = gender_dir + '-' + province_dir
            write_excel(user_lists, excel_name)
            # print(gender_dir + os.sep + province_dir)


if __name__ == '__main__':
    # user_lists = json_to_list('D:/世纪佳缘/男性用户/北京')
    # write_excel(user_lists, '男性用户-北京')
    process("D:/世纪佳缘")
